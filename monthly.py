import pandas as pd
from datetime import datetime
import os.path as osp
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import json
from googleapiclient.discovery import build
import isodate
from configparser import ConfigParser
import sys
import re
import time

# settings.iniファイルのパスを取得
# FILE = osp.join(osp.dirname(__file__), "settings.ini")
def get_base_path():
    if getattr(sys, 'frozen', False):
        return osp.dirname(sys.executable)  # 実行ファイルの場所
    return osp.dirname(__file__)  # スクリプトとして実行

FILE = osp.join(get_base_path(), "settings.ini")


# ConfigParserを使ってファイルを読み込む
settings = ConfigParser()
try:
    with open(FILE, 'r', encoding='utf-8') as f:
        settings.read_file(f)
except FileNotFoundError:
    print("Error: settings.iniファイルが見つかりません。")
    sys.exit(1)

# 環境変数からAPIキーとチャンネルIDを取得
API_KEY = settings["entity"].get("GCP_APIKEY")
CHANNEL_IDS = settings["entity"].get("CHANNEL_IDS", "").split(',')

if not API_KEY or not CHANNEL_IDS:
    print("Error: APIキーまたはチャンネルIDが取得できませんでした。'.ini'ファイルを確認してください。")
    exit(1)
    

def parse_duration(duration):
    """ISO 8601形式のDurationを秒に変換する"""
    return int(isodate.parse_duration(duration).total_seconds())

def load_cached_data(filename):
    """キャッシュデータを読み込む"""
    if osp.exists(filename):
        with open(filename, 'r') as file:
            try:
                data = json.load(file)
            except json.JSONDecodeError:
                data = {}
    else:
        data = {}
    
    # デフォルトの構造を設定
    if 'meta' not in data:
        data['meta'] = {"last_fetched_date": None}
    if 'videos' not in data:
        data['videos'] = {}
    
    return data

def save_cached_data(filename, data):
    """キャッシュデータを保存する"""
    # 動画データを公開日時でソート
    sorted_videos = dict(sorted(data['videos'].items(), key=lambda item: item[1]['publishedAt']))
    data['videos'] = sorted_videos
    
    with open(filename, 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=2)

def get_upload_playlist_id(youtube, channel_id):
    request = youtube.channels().list(
        part="contentDetails",
        id=channel_id
    )
    response = request.execute()
    uploads_id = response['items'][0]['contentDetails']['relatedPlaylists']['uploads']
    return uploads_id

    
def get_channel_info(youtube, channel_id):
    """チャンネル情報を取得する関数"""
    try:
        request = youtube.channels().list(
            part="statistics",
            id=channel_id
        )
        response = request.execute()
        return response['items'][0]['statistics']
    except Exception as e:
        print(f"Error fetching channel info: {e}")
        return None


def get_channel_name(youtube, channel_id):
    """チャンネル名を取得する関数"""
    try:
        request = youtube.channels().list(
            part="snippet",
            id=channel_id
        )
        response = request.execute()
        return response['items'][0]['snippet']['title']
    except Exception as e:
        print(f"Error fetching channel name: {e}")
        return None


def get_new_videos_from_playlist(youtube, upload_playlist_id, last_fetched_date=None):
    video_ids = []
    try:
        request = youtube.playlistItems().list(
            part="snippet,contentDetails",
            playlistId=upload_playlist_id,
            maxResults=50
        )
        while request:
            response = request.execute()
            for item in response['items']:
                published_at = item['snippet']['publishedAt']
                if last_fetched_date and published_at <= last_fetched_date:
                    return video_ids  # それ以降は取得しない
                video_id = item['contentDetails']['videoId']
                video_ids.append((video_id, published_at))
            request = youtube.playlistItems().list_next(request, response)
    except Exception as e:
        print(f"Error fetching new videos from playlist: {e}")
    return video_ids


def get_video_details(youtube, video_ids):
    """動画の詳細情報を取得する関数"""
    video_details = []
    try:
        for i in range(0, len(video_ids), 50):
            batch_ids = [vid for vid, _ in video_ids[i:i+50]]
            request = youtube.videos().list(
                part="contentDetails,statistics,snippet",
                id=','.join(batch_ids)
            )
            response = request.execute()
            video_details.extend(response['items'])
    except Exception as e:
        print(f"Error fetching video details: {e}")
    return video_details

def sanitize_filename(name):
    safe_name = re.sub(r'[\\/*?:[\]]', '_', name)
    return safe_name[:31]

def update_excel_with_all_data_to_sheet(wb, youtube, channel_id, cache_filename, channel_name):

    """Excelファイルを更新し、すべてのデータを含める"""
    excel_filename = 'monthly_channel_statistics.xlsx'
    if osp.exists(excel_filename):
        try:

            existing_df = pd.read_excel(excel_filename, sheet_name=sanitize_filename(channel_name if channel_name else channel_id))
        except Exception:
            existing_df = pd.DataFrame()
    else:
        existing_df = pd.DataFrame()

    
    try:
        # 新しい動画データの取得と集計
        sheet_name = sanitize_filename(channel_name if channel_name else channel_id)
        cached_data = load_cached_data(cache_filename)
        video_cache = cached_data.get('videos', {})
        last_fetched_date = cached_data['meta'].get('last_fetched_date', None)
        upload_playlist_id = get_upload_playlist_id(youtube, channel_id)
        new_video_ids = get_new_videos_from_playlist(youtube, upload_playlist_id, last_fetched_date)


        if new_video_ids:
            cached_data['meta']['last_fetched_date'] = max(published_at for _, published_at in new_video_ids)
            video_details = get_video_details(youtube, new_video_ids)
            for video in video_details:
                video_id = video['id']
                duration = video.get('contentDetails', {}).get('duration', 'PT0S')
                # 新しい動画情報を既存のキャッシュに追加
                if video_id not in video_cache:  # すでにキャッシュにない場合のみ追加
                    video_cache[video_id] = {
                        'duration': duration,
                        'viewCount': int(video['statistics'].get('viewCount', 0)),
                        'publishedAt': video['snippet']['publishedAt']
                    }
            cached_data['videos'] = video_cache
            save_cached_data(cache_filename, cached_data)

        # チャンネル情報取得
        channel_stats = get_channel_info(youtube, channel_id)
        if not channel_stats:
            print("Failed to retrieve channel statistics.")
            return
        subscriber_count = int(channel_stats['subscriberCount'])

        # データ集計
        data = {}
        previous_subscriber_count = None
        for vid, info in video_cache.items():
            month_str = datetime.strptime(info['publishedAt'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m')
            duration_sec = parse_duration(info['duration'])
            views = info['viewCount']
            if month_str not in data:
                data[month_str] = {
                    '登録者数': subscriber_count,
                    '獲得登録者数': 0,
                    '長編投稿数': 0,
                    'ショートの視聴者数': 0,
                    '動画の視聴数(総再生回数)': 0
                }
                
                
            if 180 <= duration_sec <= 3600:
                data[month_str]['長編投稿数'] += 1
            elif duration_sec < 180:
                data[month_str]['ショートの視聴者数'] += views
            data[month_str]['動画の視聴数(総再生回数)'] += views

        # 新しいデータフレームの作成
        new_df = pd.DataFrame([{
            'month': k,
            '登録者数': v['登録者数'],
            '獲得登録者数': v['獲得登録者数'],
            '長編投稿数': v['長編投稿数'],
            'ショートの視聴者数': v['ショートの視聴者数'],
            '動画の視聴数(総再生回数)': v['動画の視聴数(総再生回数)']
        } for k, v in data.items()])
        new_df['獲得登録者数'] = new_df['登録者数'].diff().fillna(0).astype(int)
        new_df['month'] = pd.to_datetime(new_df['month'], format='%Y-%m')

        # データの結合と重複削除
        combined_df = pd.concat([existing_df.dropna(how='all'), new_df.dropna(how='all')], ignore_index=True)
        combined_df['month'] = pd.to_datetime(combined_df['month'], errors='coerce')
        combined_df = combined_df.dropna(subset=['month'])  # NaTを除外
        combined_df = combined_df.drop_duplicates(subset='month')  # 同じ月の重複を除去
        combined_df.sort_values('month', inplace=True)
        combined_df.reset_index(drop=True, inplace=True)

        # 月を文字列で持つ列を追加（表示・比較用）
        combined_df['month_str'] = combined_df['month'].dt.strftime('%Y-%m')


        # 前月比の計算と更新
        for col in ['登録者数', '獲得登録者数', '長編投稿数', 'ショートの視聴者数', '動画の視聴数(総再生回数)']:
            if not combined_df[col].isna().all():
                increase = combined_df[col].diff().fillna(0).infer_objects(copy=False)
                previous_values = combined_df[col].shift(1).replace(0, pd.NA).infer_objects(copy=False)
                increase_percent = (increase / previous_values).fillna(pd.NA) * 100
                increase_percent_str = increase_percent.apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "N/A")
                combined_df[f'{col}の増加'] = [
                    f"{int(inc)} ({pct})" if pd.notna(inc) else "N/A"
                    for inc, pct in zip(increase, increase_percent_str)
                ]

        # Excelファイルへの書き出し
        ws = wb.create_sheet(title=sheet_name)
        ws['A1'] = '項目'

        metrics = ['登録者数', '獲得登録者数', '長編投稿数', 'ショートの視聴者数', '動画の視聴数(総再生回数)']
        months = combined_df['month_str'].drop_duplicates().tolist()

        # ヘッダー作成
        for col_idx, month in enumerate(months, start=2):
            ws.cell(row=1, column=col_idx).value = month
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

        # 各メトリクス行の作成
        for row_idx, metric in enumerate(metrics, start=2):
            ws.cell(row=row_idx, column=1).value = metric
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

            for col_idx, month in enumerate(months, start=2):
                value = combined_df.loc[combined_df['month_str'] == month, metric].values[0] if not combined_df.loc[combined_df['month_str'] == month, metric].empty else None
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.number_format = '#,##0'  # カンマ区切り
                cell.alignment = Alignment(horizontal="right")

        # 各月の前月比列を追加
        for i, metric in enumerate(metrics):
            row_idx = len(metrics) + 2 + i
            ws.cell(row=row_idx, column=1).value = f"{metric}の増加"
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

            for col_idx, month in enumerate(months, start=2):
                delta = combined_df.loc[combined_df['month_str'] == month, f'{metric}の増加']

                value = delta.values[0] if not delta.empty else "N/A"
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="right")

                # 文字色設定ロジック
                if isinstance(value, str) and value != "N/A":
                    val_part = value.split()[0]
                    if val_part.startswith('-'):
                        cell.font = Font(color="FF0000")  # 赤
                    elif val_part == "0" or ("N/A" in val_part):
                        cell.font = Font(color="000000")  # 黒
                    else:
                        cell.font = Font(color="0000FF")  # 青

        # 列幅を自動調整
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        # 保存
    except Exception as e:
        print(f"Error updating Excel file: {e}")
        
def main():
    try:
        # YouTube APIクライアントの初期化
        youtube = build('youtube', 'v3', developerKey=API_KEY)

        # Excelファイルを新しくつくる
        from openpyxl import Workbook
        excel_filename = 'monthly_channel_statistics.xlsx'
        wb = Workbook()
        wb.remove(wb.active)  # はじめの空のシートを削除


        # チャンネルIDを全部読む
        channel_ids = [cid.strip() for cid in settings["entity"].get("CHANNEL_IDS", "").split(",") if cid.strip()]

        # 1つずつチャンネルを読み込む
        for channel_id in channel_ids:
            channel_id = channel_id.strip()
            print(f"チャンネル {channel_id} を読み込み中...")

            channel_name = get_channel_name(youtube, channel_id)
            if channel_name:
                print(f"チャンネル名: {channel_name}")
            

            cache_filename = f"video_cache_{sanitize_filename(channel_name)}.json"
            update_excel_with_all_data_to_sheet(wb, youtube, channel_id, cache_filename, channel_name)

        # Excelファイルを保存！
        excel_filename = 'monthly_channel_statistics.xlsx'
        wb.save(excel_filename)
        print(f"\n✅ 完了！Excelを保存しました: {excel_filename}")

    except Exception as e:
        print(f"エラーが起きました: {e}")
        input("エンターで画面を閉じます。")

    else:
        time.sleep(5)
        sys.exit(0)    # プログラムを終了して画面を閉じる


if __name__ == "__main__":
    main()
